import io

from openpyxl import load_workbook

from app.core.exceptions import ValidationException
from app.processing.parsers.base_parser import BaseParser


class ExcelParser(BaseParser):
    """Extracts text content from Excel (.xlsx) files, sheet by sheet."""

    def extract_text(self, file_bytes: bytes) -> str:
        try:
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception as exc:
            raise ValidationException("Could not read Excel file — it may be corrupted") from exc

        sheets_text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_lines = [f"--- Sheet: {sheet_name} ---"]

            for row in sheet.iter_rows(values_only=True):
                cell_values = [str(cell) if cell is not None else "" for cell in row]
                if any(cell_values):
                    sheet_lines.append(" | ".join(cell_values))

            if len(sheet_lines) > 1:
                sheets_text.append("\n".join(sheet_lines))

        return "\n\n".join(sheets_text).strip()